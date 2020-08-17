import random
import xlrd
import numpy as np
import pandas as pd
from deap import algorithms, base, creator, tools
import multiprocessing
from time import time
from openpyxl import load_workbook
import matplotlib.pyplot as plt

from NuevaLectura import lectura_datos_excel

################################################################################
#FUNCIÓN DE CONSTRUCCIÓN DE INDIVIDUOS
def individual_creator(n):
    aux=0
    individual=[]
    for i in range(0,n):
        choice=get_choice()
        if choice==1:
            aux+=1
            if aux>(n*2/20):
                choice=0
        individual.append(choice)
        return individual



################################################################################
#FUNCIÓN DE MUTACIÓN
def mutation(population):
    mut=0
    return mut

################################################################################
#FUNCIÓN DE CROSSOVER
def mate(ind1, ind2):
    #Función de crossover en dos puntos. Funcionará igual que cxTwoPoint solo que
    #se comprobará si los individuos creados son válidos
    #:param ind1: Primer individuo que participa en el crossover
    #:param ind2: Segundo individuo que participa en el  crossover.
    #:returns: Tupla de 2 individuos
    size = min(len(ind1), len(ind2))
    #cxpoint1 = random.randint(1, size)
    #cxpoint2 = random.randint(1, size - 1)
    cxpoint1=random.randrange(0, size-20+1, 20)
    cxpoint2 = random.randrange(19, size, 20)
    #Comprueba si el punto 2 es mayor que el 1
    if cxpoint2 >= cxpoint1:
        cxpoint2 += 1
    else:  # Intercambia uno por otro
        cxpoint1, cxpoint2 = cxpoint2, cxpoint1

    ind1[cxpoint1:cxpoint2], ind2[cxpoint1:cxpoint2] \
        = ind2[cxpoint1:cxpoint2], ind1[cxpoint1:cxpoint2]

    return ind1, ind2

################################################################################
#FUNCIÓN PARA GENERAR EL INDIVIDUO
def get_choice():
    f=2/20
    if random.random() <=f:
        return 1
    else:
        return 0

#FUNCIÓN QUE CALCULA LA DIFERENCIA ENTRE EL DÍA CON MÁS HORAS Y EL DÍA CON MENOS
#Se llama desde la función evaluator
def dif_curso(individual, tam_curso, inicio):
    j = inicio
    #print(inicio)
    sum_monday = 0
    sum_tuesday = 0
    sum_wednesday = 0
    sum_thursday = 0
    sum_friday = 0
    #Multiplicamos el tam del curso por 20 (bloques temporales a la semana)
    #print("FINAL BUCLE: ", tam_curso*20+inicio)
    while j < (tam_curso*20+inicio):
        sum_monday+=sum(individual[j:j+4])
        #print("sum monday", sum_monday)
        sum_tuesday += sum(individual[j+4:j + 8])
        sum_wednesday += sum(individual[j + 8:j + 12])
        sum_thursday +=sum(individual[j + 12:j + 16])
        sum_friday +=sum(individual[j + 16:j + 20])
        j+=20
    minimo = min([sum_monday, sum_tuesday, sum_wednesday, sum_thursday, sum_friday])
    maximo = max([sum_monday, sum_tuesday, sum_wednesday, sum_thursday, sum_friday])
    #print("maximo ", maximo)
    #print("minimo ", minimo)
    result = abs(maximo - minimo)
    #print("INICIO ", inicio)
    #print("TAM CURSO:",tam_curso)
    #print("INDIVIDUO: ", individual)
    #print("RESULT: ", result)
    #print("DIFERENCIA CURSO: ", result)
    return result

#################################################################################
#FUNCIÓN QUE EVALÚA POR CURSOS
# Esta función se utiliza si se escoge el método datos_cuatrimestres
# def evaluator_cursos(individual):
#     v=[]
#     if cuatrimestre=="Primer cuatrimestre":
#         v.append(dif_curso(individual, grupos_11, 0))
#         v.append(dif_curso(individual, grupos_21, grupos_11*20))
#         v.append(dif_curso(individual, grupos_31, grupos_21*20+grupos_11*20))
#         v.append(dif_curso(individual, grupos_41, grupos_31*20+grupos_21*20+grupos_11*20))
#     else:
#         v.append(dif_curso(individual, grupos_12, 0))
#         v.append(dif_curso(individual, grupos_22, grupos_12*20))
#         v.append(dif_curso(individual, grupos_32, grupos_22*20+grupos_12*20))
#         v.append(dif_curso(individual, grupos_42, grupos_32*20+grupos_22*20+grupos_12*20))
#
#     minimo = min(v)
#     maximo = max(v)
#     result = abs(maximo - minimo)
#     result =result+ 10*check_feasibility(individual)
#     #print("Result: ", result)
#     return (result, )

#FUNCIÓN QUE EVALÚA POR CURSOS E IDIOMA
# Esta función se utiliza si se escoge el método datos_curso
def evaluator2(individual):

    keys=dict_ev.keys()
    j=0
    result=0
    for k in keys:
        #print("RESULT ANTES: ", result)
        #print("Key: ", k)
        #print("Len: ", len(dict_ev[k]))
        #print("j: ", j)
        result+=dif_curso(individual, len(dict_ev[k]), j)
        #print("RESULT DESPUÉS: ", result)
        j+=(len(dict_ev[k])*20)
    #print("suma resultado: ", result)
    #print("FEASIBILITY: ", check_feasibility(individual))
    restricciones=check_feasibility(individual)
    vector_restricciones.append(restricciones)
    #if restricciones==0:
    #    print(" %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%% RESTRICCIONES = 0 %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%")

    result =result+ 10*restricciones
    #print("RESULTADO FINAL: ", result)
    return (result, )

#################################################################################
#FUNCIÓN ORIGINAL
#Evalúa el individuo sin separar por cursos. Entero.
def evaluator(individual):
    #Aquí se debe calcular la diferencia entre el día que más horas lectivas posee
    #y el día que menos
    l=len(individual)
    result=0
    j=0
    sum_monday=0
    sum_tuesday=0
    sum_wednesday=0
    sum_thursday=0
    sum_friday=0
    #Se recorre el vector total.
    #individual es un vector que contiene la info de los grupo-asignatura por bloques
    #de forma consecutiva. Es decir, de 0-19 toda la info del Grupo-Asign1, de 20-30 Grupo-Asing2, etc
    #Además, el vector correspondiente al grupo-asign está la info de todos los bloques horarios.
    #Sabemos que cada día tiene 4 bloques posibles, luego para saber las horas por cada día serían los
    #bloques 0-3 para el lunes, 4-7 para el martes, 8-11 para el miércoles, etc.
    while j < l:
        #for i in range(j, j+4):
        #    sum_monday=sum_monday+individual[i]
        sum_monday+=sum(individual[j:j+4])
        #for i in range(j+4, j+8):
         #   sum_tuesday=sum_tuesday+individual[i]
        sum_tuesday += sum(individual[j+4:j + 8])
        #for i in range(j + 8, j + 12):
         #   sum_wednesday = sum_wednesday + individual[i]
        sum_wednesday += sum(individual[j + 8:j + 12])
        #for i in range(j + 12, j + 16):
            #sum_thursday = sum_thursday + individual[i]
        sum_thursday +=sum(individual[j + 12:j + 16])
        #for i in range(j + 16, j + 20):
            #sum_friday = sum_friday + individual[i]
        sum_friday +=sum(individual[j + 16:j + 20])
        j+=20
    #print("Monday: ", sum_monday)
    #print("Tuesday: ", sum_tuesday)
    #print("Wednesday: ", sum_wednesday)
    #print("Thursday: ", sum_thursday)
    #print("Friday: ", sum_friday)

    minimo=min([sum_monday, sum_tuesday, sum_wednesday, sum_thursday, sum_friday])
    maximo = max([sum_monday, sum_tuesday, sum_wednesday, sum_thursday, sum_friday])
    result=abs(maximo-minimo)
    #if not check_feasibility(individual):
    #    result+=2000
    result =result+ 10*check_feasibility(individual)
    #print("Result: ", result)
    return (result, )

#print("FUNCIÓN EVAL: ", evaluator(list([1, 1, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0])))

##################################################################################
#FUNCIÓN QUE ALMACENA LOS DATOS DE ENTRADA EN UN DICCIONARIO
#DICT [ "Cuatrimestre"]
# def datos_cuatrimestres():
#     grupos_11 = 0
#     grupos_21 = 0
#     grupos_31 = 0
#     grupos_41 = 0
#     grupos_12 = 0
#     grupos_22 = 0
#     grupos_32 = 0
#     grupos_42 = 0
#     for row in range(2,sheet.nrows):
#         curso = sheet.cell(row, 2).value
#         cuatri = str(sheet.cell(row, 3).value)
#         c = 0
#         if int(cuatri.find("Primer")) >= 0:
#             c = int(1)
#         else:
#             c = int(2)
#         opt = str(sheet.cell(row, 4).value)
#         if len(opt) == 0:
#             opt = ""
#         else:
#             opt = "-" + opt
#         if (curso != "X"):
#             curso = int(curso)
#         else:
#             curso=int(3)
#
#         if cuatri=="Primer cuatrimestre":
#             # Columna 5: idioma ES y magistral. Columna 6: idioma ES y laboratorio.
#             if int(sheet.cell(row, 5).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 5).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "ES-" + str(curso) + str(c) + opt
#                     vector_primer.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 14).value)
#             if int(sheet.cell(row, 6).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 6).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "ES-" + str(curso) + str(
#                         c) + opt
#                     vector_primer.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 15).value)
#             # Columna 7: idioma EU y magistral. Columna 8: idioma EU y laboratorio.
#             if int(sheet.cell(row, 7).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 7).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EU-" + str(curso) + str(c) + opt
#                     vector_primer.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 14).value)
#             if int(sheet.cell(row, 8).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 8).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EU-" + str(curso) + str(
#                         c) + opt
#                     vector_primer.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 15).value)
#             # Columna 9: idioma EN y magistral. Columna 10: idioma EN y laboratorio.
#             if int(sheet.cell(row, 9).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 9).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EN-" + str(curso) + str(c) + opt
#                     vector_primer.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 14).value)
#             if int(sheet.cell(row, 10).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 10).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EN-" + str(curso) + str(
#                         c) + opt
#                     vector_primer.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 15).value)
#
#         else:
#             # Columna 5: idioma ES y magistral. Columna 6: idioma ES y laboratorio.
#             if int(sheet.cell(row, 5).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 5).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "ES-" + str(curso) + str(c) + opt
#                     vector_segundo.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 14).value)
#             if int(sheet.cell(row, 6).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 6).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "ES-" + str(curso) + str(
#                         c) + opt
#                     vector_segundo.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 15).value)
#             # Columna 7: idioma EU y magistral. Columna 8: idioma EU y laboratorio.
#             if int(sheet.cell(row, 7).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 7).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EU-" + str(curso) + str(c) + opt
#                     vector_segundo.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 14).value)
#             if int(sheet.cell(row, 8).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 8).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EU-" + str(curso) + str(
#                         c) + opt
#                     vector_segundo.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 15).value)
#             # Columna 9: idioma EN y magistral. Columna 10: idioma EN y laboratorio.
#             if int(sheet.cell(row, 9).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 9).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EN-" + str(curso) + str(c) + opt
#                     vector_segundo.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 14).value)
#             if int(sheet.cell(row, 10).value) > 0:
#                 for i in range(1, int(sheet.cell(row, 10).value) + 1):
#                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EN-" + str(curso) + str(
#                         c) + opt
#                     vector_segundo.append(codigo)
#                     dict_horassemanales[codigo] = int(sheet.cell(row, 15).value)
#
#     dict_asignaturas["Primer cuatrimestre"]=vector_primer
#     dict_asignaturas["Segundo cuatrimestre"]=vector_segundo
#
#     for item in vector_primer:
#         if item.split("-")[2]=='11':
#             grupos_11 +=1
#         elif item.split("-")[2]=='21':
#             grupos_21 += 1
#         elif item.split("-")[2] == '31':
#             grupos_31 += 1
#         else:
#             grupos_41 += 1
#
#     for item in vector_segundo:
#         if item.split("-")[2]=='12': grupos_12 += 1
#         elif item.split("-")[2] == '22': grupos_22 += 1
#         elif item.split("-")[2] == '32': grupos_32 += 1
#         else: grupos_42+=1
#
#     return (grupos_11, grupos_21, grupos_31, grupos_41, grupos_21, grupos_22, grupos_32, grupos_42)

#################################################################################
#FUNCIÓN QUE ALMACENA LOS DATOS DE ENTRADA EN UN DICCIONARIO
#DICT [ "Curso", "Cuatrimestre", "Idioma"]
def datos_curso(file):
     read_file=xlrd.open_workbook(file)
     sheet=read_file.sheet_by_name("Asignaturas y Grupos")
     curso_aux=int(1)
     cuatri_aux="Primer cuatrimestre"
     vector_es=[]
     vector_eu=[]
     vector_en=[]
     for row in range(2,sheet.nrows):
         #Código de asignatura posición 0
         curso=sheet.cell(row,2).value
         cuatri=str(sheet.cell(row,3).value)
         c=0
         if int(cuatri.find("Primer") )>=0:
             c = int(1)
         else:
             c = int(2)
         opt=str(sheet.cell(row,4).value)
         if len(opt)==0:
             opt = ""
         else:
             opt = "-" + opt
         if (curso != "X"): curso= int(curso)
         #if (curso == "X"): curso = int(3)
         #curso=int(curso)
         #Si coinciden curso y cuatri de la fila, seguimos metiendo valores en el vector
         #print("CURSO: ", curso)
         #print("CUATRI: ", cuatri)
         if curso==curso_aux and cuatri==cuatri_aux:

             #Columna 5: idioma ES y magistral. Columna 6: idioma ES y laboratorio.
             if int(sheet.cell(row, 5).value)>0:
                 for i in range(1, int(sheet.cell(row, 5).value)+1):
                     codigo=str(int(sheet.cell(row,0).value))+"-M"+str(i).zfill(2)+"ES-"+str(curso)+str(c)+opt
                     vector_es.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")]=int(sheet.cell(row,14).value)
             if int(sheet.cell(row, 6).value)>0:
                 for i in range(1, int(sheet.cell(row, 6).value)+1):
                     codigo=str(int(sheet.cell(row,0).value))+"-GL"+str(i).zfill(2)+"ES-"+str(curso)+str(c)+opt
                     vector_es.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 15).value)
             # Columna 7: idioma EU y magistral. Columna 8: idioma EU y laboratorio.
             if int(sheet.cell(row, 7).value) > 0:
                 for i in range(1, int(sheet.cell(row, 7).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EU-"+str(curso)+str(c)+opt
                     vector_eu.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 14).value)
             if int(sheet.cell(row, 8).value) > 0:
                 for i in range(1, int(sheet.cell(row, 8).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EU-"+str(curso)+str(c)+opt
                     vector_eu.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 15).value)
             # Columna 9: idioma EN y magistral. Columna 10: idioma EN y laboratorio.
             if int(sheet.cell(row, 9).value) > 0:
                 for i in range(1, int(sheet.cell(row, 9).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EN-"+str(curso)+str(c)+opt
                     vector_en.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 14).value)
             if int(sheet.cell(row, 10).value) > 0:
                 for i in range(1, int(sheet.cell(row, 10).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EN-"+str(curso)+str(c)+opt
                     vector_en.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 15).value)
         else:

             #Cuando deja de coincidir, significa que estamos en una fila que tiene alguno de los parámetros diferentes.
             #Guardamos la info en el diccionario correspondiente, vaciamos los vectores y metemos
             #el valor de dicha fila. Será el primer valor de la siguiente entrada del diccionario.
             if(curso_aux!="X"): curso_aux=int(curso_aux)
             #if (curso_aux == "X"): curso_aux = int(3)
             dict_asignaturas[curso_aux, cuatri_aux, "ES"]=vector_es
             dict_asignaturas[curso_aux, cuatri_aux, "EU"] = vector_eu
             dict_asignaturas[curso_aux, cuatri_aux, "EN"] = vector_en
             curso_aux=curso
             cuatri_aux=cuatri
             vector_es=[]
             vector_eu = []
             vector_en = []
             if int(sheet.cell(row, 5).value)>0:
                 for i in range(1, int(sheet.cell(row, 5).value)+1):
                     codigo=str(int(sheet.cell(row,0).value))+"-M"+str(i).zfill(2)+"ES-"+str(curso)+str(c)+opt
                     vector_es.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 14).value)
             if int(sheet.cell(row, 6).value)>0:
                 for i in range(1, int(sheet.cell(row, 6).value)+1):
                     codigo=str(int(sheet.cell(row,0).value))+"-GL"+str(i).zfill(2)+"ES-"+str(curso)+str(c)+opt
                     vector_es.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 15).value)
             if int(sheet.cell(row, 7).value) > 0:
                 for i in range(1, int(sheet.cell(row, 7).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EU-"+str(curso)+str(c)+opt
                     vector_eu.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 14).value)
             if int(sheet.cell(row, 8).value) > 0:
                 for i in range(1, int(sheet.cell(row, 8).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EU-"+str(curso)+str(c)+opt
                     vector_eu.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 15).value)
             if int(sheet.cell(row, 9).value) > 0:
                 for i in range(1, int(sheet.cell(row, 9).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-M" + str(i).zfill(2) + "EN-"+str(curso)+str(c)+opt
                     vector_en.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 14).value)
             if int(sheet.cell(row, 10).value) > 0:
                 for i in range(1, int(sheet.cell(row, 10).value) + 1):
                     codigo = str(int(sheet.cell(row, 0).value)) + "-GL" + str(i).zfill(2) + "EN-"+str(curso)+str(c)+opt
                     vector_en.append(codigo)
                     dict_horassemanales[codigo.replace("X", "3")] = int(sheet.cell(row, 15).value)
     #Escribimos la última parte del excel
     dict_asignaturas[curso_aux, cuatri_aux, "ES"] = vector_es
     dict_asignaturas[curso_aux, cuatri_aux, "EU"] = vector_eu
     dict_asignaturas[curso_aux, cuatri_aux, "EN"] = vector_en

     #print("DICCIONARIO: ", dict_asignaturas)
     #Necesario convertir el curso "X" en 3º. Así que, juntamos ambas entradas del diccionario.
     #Idioma ES, Primer cuatrimestre
     vec=dict_asignaturas[3, "Primer cuatrimestre", "ES"]+dict_asignaturas['X', "Primer cuatrimestre", "ES"]
     vec_new=[]
     [vec_new.append(s.replace("X", "3")) for s in vec]
     dict_asignaturas[3, "Primer cuatrimestre", "ES"]=vec_new
     del dict_asignaturas['X', "Primer cuatrimestre", "ES"]
     #Idioma ES, Segundo cuatrimestre
     vec = dict_asignaturas[3, "Segundo cuatrimestre", "ES"] + dict_asignaturas['X', "Segundo cuatrimestre", "ES"]
     vec_new = []
     [vec_new.append(s.replace("X", "3")) for s in vec]
     dict_asignaturas[3, "Segundo cuatrimestre", "ES"] = vec_new
     del dict_asignaturas['X', "Segundo cuatrimestre", "ES"]
     # Idioma EU, Primer cuatrimestre
     vec = dict_asignaturas[3, "Primer cuatrimestre", "EU"] + dict_asignaturas['X', "Primer cuatrimestre", "EU"]
     vec_new = []
     [vec_new.append(s.replace("X", "3")) for s in vec]
     dict_asignaturas[3, "Primer cuatrimestre", "EU"] = vec_new
     del dict_asignaturas['X', "Primer cuatrimestre", "EU"]
     # Idioma EU, Segundo cuatrimestre
     vec = dict_asignaturas[3, "Segundo cuatrimestre", "EU"] + dict_asignaturas['X', "Segundo cuatrimestre", "EU"]
     vec_new = []
     [vec_new.append(s.replace("X", "3")) for s in vec]
     dict_asignaturas[3, "Segundo cuatrimestre", "EU"] = vec_new
     del dict_asignaturas['X', "Segundo cuatrimestre", "EU"]
     # Idioma EN, Primer cuatrimestre
     vec = dict_asignaturas[3, "Primer cuatrimestre", "EN"] + dict_asignaturas['X', "Primer cuatrimestre", "EN"]
     vec_new = []
     [vec_new.append(s.replace("X", "3")) for s in vec]
     dict_asignaturas[3, "Primer cuatrimestre", "EN"] = vec_new
     del dict_asignaturas['X', "Primer cuatrimestre", "EN"]
     # Idioma EN, Segundo cuatrimestre
     vec = dict_asignaturas[3, "Segundo cuatrimestre", "EN"] + dict_asignaturas['X', "Segundo cuatrimestre", "EN"]
     vec_new = []
     [vec_new.append(s.replace("X", "3")) for s in vec]
     dict_asignaturas[3, "Segundo cuatrimestre", "EN"] = vec_new
     del dict_asignaturas['X', "Segundo cuatrimestre", "EN"]


    #Necesario modificar también el diccionario de horas semanales
    #Lo modificamos haciendo un replace cuando se introducen los datos. En caso de que exista una X,
    #la reemplazará por un 3
    #print("HORAS SEMANALES", dict_horassemanales)

#print("GRUPOS 42: ", grupos_42)
#print(dict_asignaturas["Primer cuatrimestre"])
#print(dict_asignaturas)
#print(dict_horassemanales)
#print(len(dict_horassemanales))



##################################################################################

#FUNCIÓN DE COMPROBACIÓN DE RESTRICCIONES
def check_feasibility(individual):
    # Esta función sirve para comprobar si el individuo
    # cumple o no, las restricciones del problema

    #Restricción sobre los solapamientos de las asignaturas
    ind_df = pd.DataFrame(np.reshape(individual, (tam, 20)), index=asignaturas, columns=bloques)
    matrix=np.reshape(individual, (tam, 20))
    #penalty=penalty + abs(sum(row)
           # - dict_horassemanales[ind_row]) if sum(row) != dict_horassemanales[ind_row] for ind_row, row in ind_df.iterrows()
    #print("DESPUÉS DE BUCLE OPTIMIZADO: ",penalty)

    penalty=0
    for ind_row, row in ind_df.iterrows():
        s = sum(row)
        if s != dict_horassemanales[ind_row]:
            penalty += abs(s - dict_horassemanales[ind_row])


    #Se multiplica la matriz del individuo por su transpuesta. De esta forma tendremos
    #una matriz con dimensiones asignaturas x asinaturas donde cada valor refleja el
    #número de veces que cada grupo i coincide con el grupo j
    new_matrix = matrix.dot(np.transpose(matrix))
    #Creamos un dataframe de dicha matriz
    new_df=pd.DataFrame(new_matrix, index=asignaturas, columns=asignaturas)
    #Multiplicamos ambos dataframes porque no es lo mismo multiplicar matrices que dataframes.
    #De esta forma multiplica aquellos que tienen los mismos index. Así, tendremos una matriz
    #resultado en new_df con valores únicamente para los grupos que son incompatibles.
    new_df=new_df*incomp_df
    #Obtenemos dichos valores
    values=new_df.values
    #Sumamos los valores de la diagonal inferior (sin contar la diagonal principal). Este será
    #el valor de penalización que estamos buscando.
    suma_diagonal_inf=sum(values[np.tril_indices(values.shape[0], -1)])
    penalty+=suma_diagonal_inf
    #print("SUMA DIAGONAL: ", suma_diagonal_inf)
    ###############
    # #MÉTODO LENTO
    # #penalty_lento=0
    # for col in ind_df:
    #     # print(col)
    #     # aux_df=incomp_df.loc[:, col]
    #     # Recorremos el DF por columnas y nos quedamos con aquellas filas cuyo valor es 1.
    #     condicion = ind_df[col] >= 1
    #     # print(df[condicion][col])
    #     aux_df = ind_df[condicion][col]
    #     # print(aux_df.index)
    #     # print(incomp_df)
    #     for ind in list(aux_df.index):
    #         #print(ind)
    #         list_aux = list(aux_df.index)
    #         list_aux.remove(ind)
    #         #print(list_aux)
    #         for ind_aux in list_aux:
    #             #print(ind_aux)
    #             # Buscamos en la matriz de incompatibilidades, para ver si dichos grupos se pueden solapar
    #             if incomp_df[ind_aux][ind] == 1:
    #                 penalty+=1
    #                 #penalty_lento+=1
    #                 #feasibility = False
    #                 #return feasibility
    # #print("PENALTY LENTO: ", penalty_lento)
    return penalty



################################################################################
#ALGORITMO GENÉTICO


def main(filename):

    time_inicio_algoritmo = time()
    vector_resultados=[]
    if __name__ == "__main__":
        #Consideramos la función de fitness como FitnessMin y creamos los individuales como listas
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", list, fitness=creator.FitnessMin)

        #Existen 21 grupos en el primer curso del primer cuatrimestre, en español.

        #De esta forma, creamos los individuos, estarán formados
        # por 1 o 0, en función de si se imparte clase durante ese bloque o no.
        #El tamaño del vector será de 21grupos *5 días*4 bloques/día.
        #Sería una matriz de 21x20
        IND_SIZE=tam*20

        #Ahora, necesitamos un toolbox.
        #Se generarán de forma aleatoria las posibles individuos que tendrán
        #un total de 21x20 elementos y estarán formados por enteros 0 o 1.
        toolbox = base.Toolbox()

        #toolbox.register("attr_bool", random.randint, 0, 1)
        toolbox.register("attr_bool", get_choice)
        toolbox.register("individual", tools.initRepeat, creator.Individual,
                     toolbox.attr_bool, n=IND_SIZE)
        toolbox.register("population", tools.initRepeat, list, toolbox.individual)

        #Para varios procesadores
        pool=multiprocessing.Pool(processes=3)
        toolbox.register("map", pool.map)
        #Utilizamos la función evaluator que creamos como función de evaluación
        toolbox.register("evaluate", evaluator2)
        #Añadimos la función para penalizar en caso de no cumplir las restricciones
        #toolbox.decorate("evaluate", tools.DeltaPenalty(check_feasibility, 20000))
        #toolbox.register("mate", tools.cxOnePoint)
        toolbox.register("mate", mate)
        toolbox.register("mutate", tools.mutFlipBit, indpb=0.05)
        toolbox.register("select", tools.selTournament, tournsize=3)

        # Creamos la población
        #############################################################
        #############################################################
        #############################################################
        m=10
        pop = toolbox.population(n=IND_SIZE*m)
        hof = tools.HallOfFame(20, similar=np.array_equal)
        fitnesses = list(map(toolbox.evaluate, pop))
        for ind, fit in zip(pop, fitnesses):
            ind.fitness.values = fit
            # print("FIT: ", fit)

        # CXPB = probabilidad con las que dos invidividuos se cruzan
        # MUTPB = probabilidad de mutar un individuo
        CXPB, MUTPB = 0.5, 0.2

        # Se almacenan en fits los valores de fitness de los individuos
        fits = [ind.fitness.values[0] for ind in pop]

        #fbest = np.ndarray((100, 1))
        # g=variable que cuenta el número de iteraciones realizadas
        g = 0
        # Comienza la evolución
        while min(fits) > 0 and g < 75:
            # A new generation
            g = g + 1
            print("-- Generation %i --" % g)
            # Seleccionamos la población
            offspring = toolbox.select(pop, len(pop))
            # Se clonan los individuos
            offspring = list(map(toolbox.clone, offspring))
            # Se aplican crossover y mutación
            for child1, child2 in zip(offspring[::2], offspring[1::2]):
                if random.random() < CXPB:
                    toolbox.mate(child1, child2)
                    del child1.fitness.values
                    del child2.fitness.values

            for mutant in offspring:
                if random.random() < MUTPB:
                    toolbox.mutate(mutant)
                    del mutant.fitness.values

            # Se evalúa el fitness de los nuevos individuos
            # Se reevalúan aquellos casos donde se consideró inválido el valor de fitness
            invalid_ind = [ind for ind in offspring if not ind.fitness.valid]
            fitnesses = map(toolbox.evaluate, invalid_ind)
            for ind, fit in zip(invalid_ind, fitnesses):
                ind.fitness.values = fit

            # Se reemplaza la población por la nueva generada
            # después del crossover y la mutación
            pop[:] = offspring

            #Se hace un update del HallOfFame donde se almacenará la mejor solución
            hof.update(pop)
            # Se evalúa la solución
            fits = [ind.fitness.values[0] for ind in pop]

            #Resultados fitness
            #length = len(pop)
            #mean = sum(fits) / length
            #sum2 = sum(x * x for x in fits)
            #std = abs(sum2 / length - mean ** 2) ** 0.5
            #f_order=sorted(fits)
            #if (length/2)%2==0:
             #   f_mediana=(f_order[(length/2)+1]+f_order[(length/2)+2])/2
            #else:
            #    f_mediana=f_order[(length/2)+1]
            f_cuartil25, f_mediana, f_cuartil75 = np.percentile(fits, [25, 50, 75])
            #Resultados restricciones
            #length_restr=len(vector_restricciones)
            #r_order=sorted(vector_restricciones)
            #if(length_restr/2)%2==0:
            #    r_mediana=(r_order[(length_restr/2)+1]+r_order[(length_restr/2)+2])/2
            #else:
            #    r_mediana=r_order[(length_restr/2)+1]
            r_cuartil25, r_mediana, r_cuartil75=np.percentile(vector_restricciones, [25, 50, 75])
            #restricciones=sum(x for x in vector_restricciones)/len(vector_restricciones)
            vector_resultados.append([min(fits), max(fits), f_mediana, f_cuartil25, f_cuartil75, min(vector_restricciones),
                                      max(vector_restricciones), r_mediana, r_cuartil25, r_cuartil75])
            print("  Min %s" % min(fits))
            #fbest[g] = hof[0].fitness.values
            #print("  Max %s" % max(fits))
            #print("  Avg %s" % mean)
            #print("  Std %s" % std)
            #print("  Restricciones %s" % restricciones)


        #x = list(range(0, 100))
        #logbook = tools.Logbook()
        #min_fit = logbook.select("min")

        # Dibujamos
        #plt.figure()

        #plt.semilogy(x, fbest, "b-")
        #plt.semilogy(x, min_fit, "r-")
        #plt.xlabel("Generation")
        #plt.ylabel("Fitness")
        #plt.title("blue: f-best, red: min")
        #plt.show()


        #if list(hof[0])!=None:
            # Almacenamos la mejor solución en un excel para realizar comprobaciones
        #    print(list(hof[0]))
        #    print(list(hof))
        #    df = pd.DataFrame(data=np.reshape(list(hof[0]), (tam, 20)),
        #                      columns=bloques, index=asignaturas)
        #    out_file = "solucion.xlsx"
         #   df.to_excel(out_file, index=True)

        time_final=time()
        #df = pd.DataFrame(data=vector_resultados,
         #                 columns=["Min Fitness", "Max Fitness", "Mediana Fitness","Cuartil 25 Fitness", "Cuartil 75 Fitness",
         #                          "Min Restricciones", "Max Restricciones", "Mediana Restricciones", "Cuartil 25 Restricciones", "Cuartil 75 Restricciones"])
        #df.to_excel("C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\" \
     #"MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\"+asignaturas[1]+"numpruebas_"+str(m)+".xlsx", index=False)
        print("Tiempo ejecución algoritmo: ", time_final-time_inicio_algoritmo)

        df_result = pd.DataFrame(data=vector_resultados,
                                 columns=["Min Fitness", "Max Fitness", "Mediana Fitness", "Cuartil 25 Fitness",
                                          "Cuartil 75 Fitness",
                                          "Min Restricciones", "Max Restricciones", "Mediana Restricciones",
                                          "Cuartil 25 Restricciones", "Cuartil 75 Restricciones"])


        vector_i = [[ind, evaluator2(ind), check_feasibility(ind)] for ind in list(hof)]

        df_ind = pd.DataFrame(data=vector_i, columns=["Individuo", "Fitness", "Restricciones incumplidas"])

        df_t = pd.DataFrame(data=[time_final-time_inicio_algoritmo], columns=["Tiempo ejecución"])

        #Primero creamos la primera hoja del excel
        #df_result.to_excel("C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\" \
        #                "MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\Resultados_11_Pop10.xlsx",
          #                 sheet_name="Resultados", index=False)
        df_result.to_excel(filename,
                           sheet_name="Resultados", index=False)
        #Una vez creado, lo cargamos para continuar escribiendo
        #book=load_workbook("C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\Resultados_11_Pop10.xlsx")
        #writer = pd.ExcelWriter("C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\" \
                       # "MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\Resultados_11_Pop10.xlsx", engine='openpyxl')
        book=load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        #Continuamos escribiendo
        df_ind.to_excel(writer,
                        sheet_name="Mejores individuos", index=False)
        writer.save()
        df_t.to_excel(writer,
                      sheet_name="Tiempo ejecución", index=False)
        writer.save()
        #return hof, vector_resultados, time_final-time_inicio_algoritmo


#time_inicio=time()
###################################################################################
#LEER ARCHIVOS CON DATOS DE ENTRADA
#Leemos el archivo excel (versión antigua)
#file="C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\" \
     #"MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\AsignaturasYGrupos-Dayana.xlsx"
#Versión lectura antigua
#read_file=xlrd.open_workbook(file)
#sheet=read_file.sheet_by_name("Asignaturas y Grupos")

#Leemos el archivo excel (versión nueva)
file="C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\" \
     "MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\AsignaturasGruposProfesorado-20200702-Dayana.xlsx"

dict_asignaturas={}
dict_horassemanales={}
dict_profes={}
vector_resultados=[]
vector_restricciones=[]
df_result=pd.DataFrame()
#vector_es=[]
#vector_eu=[]
#vector_en=[]

#####################  LECTURA NUEVA VERSIÓN EXCEL ###############
dict_asignaturas, dict_horassemanales, dict_profes=lectura_datos_excel(file)
##################################################

################ LECTURA ANTIGUA VERSIÓN EXCEL ##################
#datos_curso(file)
#################################################

#curso_aux=int(1)
#cuatri_aux="Primer cuatrimestre"
#vector_primer=[]
#vector_segundo=[]
#cuatrimestre="Primer cuatrimestre"
#(grupos_11, grupos_21, grupos_31, grupos_41, grupos_12, grupos_22, grupos_32, grupos_42) = (0,0,0,0,0,0,0,0)
#Ejecutamos la lectura de datos, existen dos opciones: Diccionario por cuatrimestres o por curso, idioma, cuatri
#(grupos_11, grupos_21, grupos_31, grupos_41, grupos_12, grupos_22, grupos_32, grupos_42) = datos_cuatrimestres()
#datos_curso()
#################################################################################
#VARIABLES GLOBALES
#Ejemplo cuatrimestre , 1º curso
#Caso de diccionario por cuatrimestres
#cuatrimestre="Primer cuatrimestre"
#tam=len(dict_asignaturas[cuatrimestre])
#asignaturas=dict_asignaturas[cuatrimestre]
#asignaturas.sort(key= lambda x: int(x.split("-")[2]))

#print(dict_asignaturas[3, "Primer cuatrimestre", "ES"])

#Caso de diccionario por curso, idioma, cuatrimestre
#Versión lectura antigua
#tam=len(dict_asignaturas[2, "Primer cuatrimestre", "ES"])+len(dict_asignaturas[2, "Primer cuatrimestre", "EU"])+\
    #len(dict_asignaturas[2, "Primer cuatrimestre", "EN"])
#tam=len(dict_asignaturas[1, "Primer cuatrimestre", "ES"])+len(dict_asignaturas[1, "Primer cuatrimestre", "EU"])+\
#  len(dict_asignaturas[2, "Primer cuatrimestre", "ES"])+len(dict_asignaturas[2, "Primer cuatrimestre", "EU"])+len(dict_asignaturas[2, "Primer cuatrimestre", "EN"])
    #len(dict_asignaturas[3, "Primer cuatrimestre", "ES"])+len(dict_asignaturas[3, "Primer cuatrimestre", "EU"])
#tam=len(dict_asignaturas[3, "Primer cuatrimestre", "ES"])+len(dict_asignaturas[3, "Primer cuatrimestre", "EU"])

#Versión nueva lectura
tam=len(dict_asignaturas[3, 2, "ES"])+len(dict_asignaturas[3, 2, "EU"])

#Versión antigua lectura
# dict_ev={}
# dict_ev[1, "Primer cuatrimestre", "ES"]=dict_asignaturas[1, "Primer cuatrimestre", "ES"]
# dict_ev[1, "Primer cuatrimestre", "EU"]=dict_asignaturas[1, "Primer cuatrimestre", "EU"]
# dict_ev[2, "Primer cuatrimestre", "ES"]=dict_asignaturas[2, "Primer cuatrimestre", "ES"]
# dict_ev[2, "Primer cuatrimestre", "EU"]=dict_asignaturas[2, "Primer cuatrimestre", "EU"]
# dict_ev[2, "Primer cuatrimestre", "EN"]=dict_asignaturas[2, "Primer cuatrimestre", "EN"]
#dict_ev[3, "Primer cuatrimestre", "ES"]=dict_asignaturas[3, "Primer cuatrimestre", "ES"]
#dict_ev[3, "Primer cuatrimestre", "EU"]=dict_asignaturas[3, "Primer cuatrimestre", "EU"]

#Versión nueva lectura
dict_ev={}
dict_ev[3, 2, "ES"]=dict_asignaturas[3, 2, "ES"]
dict_ev[3, 2, "EU"]=dict_asignaturas[3, 2, "EU"]

#Versión lectura antigua
#print(dict_ev)
#asignaturas=dict_asignaturas[1, "Primer cuatrimestre", "ES"]+dict_asignaturas[1, "Primer cuatrimestre", "EU"]+\
           #dict_asignaturas[2, "Primer cuatrimestre", "ES"]+dict_asignaturas[2, "Primer cuatrimestre", "EU"]+dict_asignaturas[2, "Primer cuatrimestre", "EN"]
#asignaturas = dict_asignaturas[1, "Primer cuatrimestre", "ES"]+dict_asignaturas[1, "Primer cuatrimestre", "EU"]+\
 #dict_asignaturas[2, "Primer cuatrimestre", "ES"]+dict_asignaturas[2, "Primer cuatrimestre", "EU"]+dict_asignaturas[2, "Primer cuatrimestre", "EN"]

#Versión nueva lectura
asignaturas = dict_asignaturas[3, 2, "ES"]+dict_asignaturas[3, 2, "EU"]


#print(dict_ev)
asignaturas.sort(key= lambda x: int(x.split("-")[2]))

bloques=['Lunes1', "Lunes2", "Lunes3", "Lunes4", "Martes1", "Martes2", "Martes3", "Martes4",
         "Miércoles1", "Miércoles2", "Miércoles3", "Miércoles4", "Jueves1", "Jueves2", "Jueves3",
         "Jueves4", "Viernes1", "Viernes2", "Viernes3", "Viernes4"]

#print(asignaturas)
###################################################################################
#MATRIZ DE INCOMPATIBILIDADES
#Creamos la matriz de incompatibilidades
#Podrán solaparse aquellos grupos que compartiendo curso, cuatri e idioma, sean de tipo laboratorio.
#tam=len(dict_asignaturas[1, "Primer cuatrimestre", "ES"])+len(dict_asignaturas[1, "Segundo cuatrimestre", "ES"])
#tam=len(dict_asignaturas["Primer cuatrimestre"])
matrix=np.ones([tam, tam], dtype=np.int)
#v=dict_asignaturas["Primer cuatrimestre"]
#Ordenamos el vector de asignaturas por cursos
#v.sort(key= lambda x: int(x.split("-")[2]))
#print("ORDENADO",v)

#print(v)
#incomp_df=pd.DataFrame(matrix, index=dict_asignaturas[1, "Primer cuatrimestre", "ES"], columns=dict_asignaturas[1, "Primer cuatrimestre", "ES"])
incomp_df=pd.DataFrame(matrix, index=asignaturas, columns=asignaturas)
#print(incomp_df)

#Creamos un dataframe para tener en indice_fila y col los códigos únicos de cada asignatura.
#Primero lo creamos con 1.0 y después, lo cambiamos por un 0 en los casos en los que
#se trate de dos grupos de laboratorio de diferente asignatura. Un 0 indica que son compatibles.
#También es necesario añadir la condición de compatibilidad si se trata de asignaturas de
#diferente cuatrimestre, de diferente idioma y de diferente curso.
for col in incomp_df:
    #print(col)
    curso_cuatri=int(col.split("-")[2])
    grupo_idioma=col.split("-")[1]
    idioma=grupo_idioma[len(grupo_idioma)-2]+grupo_idioma[len(grupo_idioma)-1]
    #print(curso_cuatri)
    if "GL" in col:
        for indice_fila, fila in incomp_df.iterrows():
            #print(indice_fila)
            #print(fila)
            if ("GL" in indice_fila and indice_fila!=col) or ("GA" in indice_fila and indice_fila!=col):
                incomp_df.at[indice_fila, col]=0

    #Versión nueva lectura (tipo GA existe)
    if "GA" in col:
        for indice_fila, fila in incomp_df.iterrows():
            #print(indice_fila)
            #print(fila)
            if ("GA" in indice_fila and indice_fila!=col) or ("GL" in indice_fila and indice_fila!=col) :
                incomp_df.at[indice_fila, col]=0

    #Versión lectura antigua
    #if "LC" in col:
     #   opt = col.split("-")[3]
     #   for indice_fila, fila in incomp_df.iterrows():
     #       if "LC" in indice_fila and indice_fila != col and indice_fila.split("-")[3] != opt:
     #           incomp_df.at[indice_fila, col] = 0

    #Versión nueva lectura
    if ("3C" or "IC" or "SI") in col:
         #print("COL: ", col)
         opt = col.split("-")[3]
         #print("OPT: ", opt)
         for indice_fila, fila in incomp_df.iterrows():
             if ("3C" in indice_fila and indice_fila!=col and indice_fila.split("-")[3]!=opt) or ("IC" in indice_fila and indice_fila!=col and indice_fila.split("-")[3]!=opt) or("SI" in indice_fila and indice_fila != col and indice_fila.split("-")[3] != opt):
                incomp_df.at[indice_fila, col]=0
                #print(indice_fila)


    for indice_fila, fila in incomp_df.iterrows():
        curso_cuatri_aux=int(indice_fila.split("-")[2])
        if curso_cuatri_aux!=curso_cuatri:
            incomp_df.at[indice_fila, col] = 0

    for indice_fila, fila in incomp_df.iterrows():
        grupo_idioma_aux=indice_fila.split("-")[1]
        idioma_aux = grupo_idioma_aux[len(grupo_idioma_aux)-2] + grupo_idioma_aux[len(grupo_idioma_aux)-1]
        if idioma_aux!=idioma:
            incomp_df.at[indice_fila, col] = 0

print(incomp_df.loc[:,"26212-GA01EU-32-3C"])

#print(dict_horassemanales)
#print(dict_asignaturas[1, "Primer cuatrimestre", "ES"])
################################################################################
#INFORMACIÓN SOBRE LABORATORIOS
#Leemos la información de la hoja Laboratorios
#laboratorios_df=pd.read_excel(file, sheet_name="Laboratorios", header=0, index_col=False)
laboratorios_df=pd.read_excel(file, sheet_name="LAboratorios", header=0, index_col=False)
#print(laboratorios_df)

#time_inicio_algoritmo=0
#time_final=0
###########################################################################################################
#EJECUTAMOS EL ALGORITMO GENÉTICO
###########################################################################################################
#Ejecutamos el algoritmo genético 10 veces
# i=8
# while i<10:
#     vector_resultados = []
#     vector_restricciones = []
#     #print(" ------ EJECUCIÓN NÚMERO ", i, "  ------")
#     main("C:\\Users\\tr5568\\Desktop\\DAYANA\\PERSONAL\\" \
#                        "MÁSTER INGENIERÍA COMPUTACIONAL Y SISTEMAS INTELIGENTES\\TFM\\RESULTADOS\\Resultados_11y21_Pop30_"+str(i)+".xlsx")
#     i+=1
